import math
import matplotlib.pyplot as plt
import numpy as np

PLOT = True   # Set to False to skip the graph
EXPORT_XLSX   = True

#Code fOr calculatIon of uSing a drivetrAin aka COISA
# ===============================================================
#                       C.O.I.S.A
# ================================================================

# --- Motor ---
# Works with any brushed DC motor. 
# Common motors: 1020 coreless, 720 coreless, N20 gearmotor, 816 brushed
MOTOR_KV            = 12000   # RPM per volt. Typical 1020 coreless range: 8000-18000
MOTOR_VOLTAGE       = 3.7     # Volts. 1S lipo=3.7, 2S lipo=7.4, NiMH 4x=4.8
MOTOR_STALL_TORQUE  = 1.4     # mNm. Typical 1020: 0.8-1.5 mNm
MOTOR_STALL_CURRENT = 1.8     # Amps per motor

# --- Your current gear setup ---
MY_PINION_TEETH     = 11      # Teeth on the motor shaft gear (pinion)
MY_WHEEL_TEETH      = 70      # Teeth on the wheel gear
GEAR_MODULE         = 0.3     # Tooth size in mm. Use 0.3 or 0.4 for this motor scale
GEAR_EFFICIENCY     = 0.85    # Power lost in mesh. Spur gears: 0.80-0.92

# --- Wheel ---
# If your wheel IS the gear (teeth on the rim), set WHEEL_DIAMETER_MM = None
# and the diameter will be calculated automatically from module x teeth.
# Otherwise enter your actual wheel diameter in mm.
WHEEL_DIAMETER_MM   = None

# --- Robot ---
ROBOT_MASS_G        = 150      # Total robot weight in grams (weigh it with the battery)
FRICTION_COEFF      = 0.45    # Wheel-to-floor grip. Rubber on smooth floor: 0.4-0.6
                              # Raw printed nylon gear on floor: 0.2-0.3
NUM_DRIVE_WHEELS    = 4       # Total wheels in contact with the ground (2 or 4)
NUM_MOTORS          = 2       # Motors driving those wheels. Typical: 2 (one per side), or 4 (one per wheel)

# --- Target speed (optional) ---
# PWM (Pulse Width Modulation) (AI explanation) is how your microcontroller controls motor speed
# without changing the battery voltage. Instead of lowering the voltage to slow
# the motor down, PWM sends full voltage but switches it ON and OFF thousands of
# times per second. The percentage is how long it stays ON each cycle — 25% means
# on for a quarter of each cycle, 50% means half, 100% means always on. In your
# line follower code this is the value sent to the motor driver (e.g.
# analogWrite(motorPin, 128) on Arduino = ~50% PWM). Your PID loop adjusts this
# value in real time to keep the robot on the line. The table below shows real
# speed and torque at each PWM level so you tune with actual numbers, not guesswork.
TARGET_SPEED_MS     = None    # e.g. 2.5 for 2.5 m/s. None = let torque decide

# --- PWM operating points to evaluate ---
PWM_POINTS          = [25, 50, 75, 100]   # percent

# --- Ratio search range (for the ideal ratio finder) ---
PINION_RANGE        = range(8, 16)        # min to max pinion teeth to test
WHEEL_RANGE         = range(30, 91)       # min to max wheel teeth to test
MIN_CENTER_DIST_MM  = 7.0     # minimum printable center distance (structural limit)
MAX_WHEEL_DIAM_MM   = 40.0    # maximum wheel diameter that fits your chassis


#  INPUT VALIDATION
# ================================================================
errors = []
#I know i can make a list of tuple but ahhh it works
if MOTOR_KV <= 0:
    errors.append("MOTOR_KV must be greater than 0")
if MOTOR_VOLTAGE <= 0:
    errors.append("MOTOR_VOLTAGE must be greater than 0")
if MOTOR_STALL_TORQUE <= 0:
    errors.append("MOTOR_STALL_TORQUE must be greater than 0")
if MOTOR_STALL_CURRENT <= 0:
    errors.append("MOTOR_STALL_CURRENT must be greater than 0")
if MY_PINION_TEETH <= 0:
    errors.append("MY_PINION_TEETH must be greater than 0")
if MY_WHEEL_TEETH <= 0:
    errors.append("MY_WHEEL_TEETH must be greater than 0")
if GEAR_MODULE <= 0:
    errors.append("GEAR_MODULE must be greater than 0")
if not (0 < GEAR_EFFICIENCY <= 1):
    errors.append("GEAR_EFFICIENCY must be between 0 and 1 (e.g. 0.85)")
if WHEEL_DIAMETER_MM is not None and WHEEL_DIAMETER_MM <= 0:
    errors.append("WHEEL_DIAMETER_MM must be greater than 0 or None")
if ROBOT_MASS_G <= 0:
    errors.append("ROBOT_MASS_G must be greater than 0")
if not (0 < FRICTION_COEFF <= 1):
    errors.append("FRICTION_COEFF must be between 0 and 1 (e.g. 0.45)")
if NUM_DRIVE_WHEELS <= 0:
    errors.append("NUM_DRIVE_WHEELS must be greater than 0")
if NUM_MOTORS <= 0:
    errors.append("NUM_MOTORS must be greater than 0")
if TARGET_SPEED_MS is not None and TARGET_SPEED_MS <= 0:
    errors.append("TARGET_SPEED_MS must be greater than 0 or None")
if MIN_CENTER_DIST_MM <= 0:
    errors.append("MIN_CENTER_DIST_MM must be greater than 0")
if MAX_WHEEL_DIAM_MM <= 0:
    errors.append("MAX_WHEEL_DIAM_MM must be greater than 0")

if errors:
    print()
    print("  CONFIGURATION ERROR — fix the following before running:")
    print()
    for e in errors:
        print(f"    ✗  {e}")
    print()
    exit(1)



#  CALCULATIONS — do not edit below this line
# ================================================================

def wheel_diameter(pinion, wheel):
    if pinion == MY_PINION_TEETH and wheel == MY_WHEEL_TEETH and WHEEL_DIAMETER_MM is not None:
        return WHEEL_DIAMETER_MM
    return GEAR_MODULE * wheel

def center_distance(pinion, wheel):
    return (GEAR_MODULE * pinion + GEAR_MODULE * wheel) / 2

def evaluate(pinion, wheel, pwm_pct=50):
    ratio      = wheel / pinion
    pwm        = pwm_pct / 100
    w_diam     = wheel_diameter(pinion, wheel)
    w_rad      = (w_diam / 2) / 1000
    c_dist     = center_distance(pinion, wheel)

    motor_rpm  = MOTOR_KV * MOTOR_VOLTAGE * pwm
    wheel_rpm  = motor_rpm / ratio
    speed_ms   = (wheel_rpm / 60) * 2 * math.pi * w_rad

    w_torque   = MOTOR_STALL_TORQUE * ratio * GEAR_EFFICIENCY * pwm
    mass_kg    = ROBOT_MASS_G / 1000
    req_torque = (mass_kg * 9.81 * FRICTION_COEFF * w_rad * 1000) / NUM_DRIVE_WHEELS
    t_margin   = ((w_torque - req_torque) / w_torque) * 100 if w_torque > 0 else -999

    drive_force   = w_torque / 1000 / w_rad
    friction_lim  = mass_kg * 9.81 * FRICTION_COEFF
    slip          = drive_force > friction_lim

    current_total = MOTOR_STALL_CURRENT * pwm * NUM_MOTORS

    return {
        "ratio":        round(ratio, 2),
        "motor_rpm":    round(motor_rpm),
        "wheel_rpm":    round(wheel_rpm),
        "speed_ms":     round(speed_ms, 2),
        "speed_kmh":    round(speed_ms * 3.6, 2),
        "wheel_torque": round(w_torque, 3),
        "req_torque":   round(req_torque, 3),
        "t_margin":     round(t_margin, 1),
        "current":      round(current_total, 2),
        "slip":         slip,
        "w_diam":       round(w_diam, 2),
        "c_dist":       round(c_dist, 2),
    }

def score_ratio(pinion, wheel):
    # Scores a gear combination from 0 to 100. Higher = better for this robot.
    r = evaluate(pinion, wheel, pwm_pct=50)

    # Reject mechanically impossible combinations
    if r["c_dist"] < MIN_CENTER_DIST_MM:  return -1
    if r["w_diam"] > MAX_WHEEL_DIAM_MM:   return -1

    score = 0

    # Torque margin at 50% PWM — most important factor
    # Ideal: 40-80% margin (enough headroom for acceleration bursts and cornering)
    tm = r["t_margin"]
    if tm >= 40:   score += 40
    elif tm >= 25: score += 30
    elif tm >= 10: score += 15
    elif tm >= 0:  score += 5
    else:          return -1   # cannot move the robot — disqualified

    # Speed scoring
    spd = r["speed_ms"]
    if TARGET_SPEED_MS is not None:
        # Score by proximity to target speed at 50% PWM
        diff = abs(spd - TARGET_SPEED_MS)
        if diff <= 0.3:   score += 35
        elif diff <= 0.8: score += 25
        elif diff <= 1.5: score += 10
        else:             score += 0
    else:
        # No target — reward useful speed range (0.8-5 m/s at 50% PWM)
        if 0.8 <= spd <= 5.0:   score += 35
        elif 0.5 <= spd <= 7.0: score += 20
        else:                    score += 5

    # Current draw — keep below 3.5A total at 50% PWM
    cur = r["current"]
    if cur <= 2.0:   score += 15
    elif cur <= 3.0: score += 10
    elif cur <= 3.5: score += 5
    else:            score += 0

    # Mechanical quality — reward larger center distances (easier to print reliably)
    cd = r["c_dist"]
    if cd >= 10:   score += 10
    elif cd >= 7:  score += 7
    else:          score += 3

    return score

def rating_label(t_margin):
    if t_margin >= 40:  return "EXCELLENT"
    if t_margin >= 25:  return "GOOD     "
    if t_margin >= 10:  return "MARGINAL "
    if t_margin >= 0:   return "WEAK     "
    return                      "FAIL     "

def slip_label(slip):
    return "YES — wheel will spin" if slip else "No"


#  REPORT
# ================================================================

W    = 64
SEP  = "=" * W
SEP2 = "-" * W

print()
print(SEP)
print("  LINE FOLLOWER DRIVETRAIN CALCULATOR")
print("  Official League Tool — Brushed DC Motors")
print(SEP)

# ---------- Robot profile ----------
print(f"""
  ROBOT PROFILE
  {'Motor KV':<28}: {MOTOR_KV} rpm/V
  {'Supply voltage':<28}: {MOTOR_VOLTAGE} V
  {'No-load RPM (100% PWM)':<28}: {round(MOTOR_KV * MOTOR_VOLTAGE):,} rpm
  {'Stall torque':<28}: {MOTOR_STALL_TORQUE} mNm
  {'Robot mass':<28}: {ROBOT_MASS_G} g
  {'Wheel-floor friction coeff':<28}: {FRICTION_COEFF}
  {'Target speed':<28}: {"auto (torque-based)" if TARGET_SPEED_MS is None else f"{TARGET_SPEED_MS} m/s"}
""")

# ---------- Current ratio ----------
MY_RATIO = MY_WHEEL_TEETH / MY_PINION_TEETH
MY_WDIAM = wheel_diameter(MY_PINION_TEETH, MY_WHEEL_TEETH)
MY_CDIST = center_distance(MY_PINION_TEETH, MY_WHEEL_TEETH)

print(SEP)
print(f"  YOUR RATIO: {MY_PINION_TEETH}:{MY_WHEEL_TEETH}  ({MY_RATIO:.2f}:1)")
print(SEP)
print(f"  {'Module':<28}: {GEAR_MODULE} mm")
print(f"  {'Pinion pitch diameter':<28}: {round(GEAR_MODULE * MY_PINION_TEETH, 2)} mm")
print(f"  {'Wheel pitch diameter':<28}: {round(MY_WDIAM, 2)} mm")
print(f"  {'Center distance (CAD)':<28}: {MY_CDIST} mm  <- motor shaft to wheel axle in your design")
print(f"  {'Wheel circumference':<28}: {round(math.pi * MY_WDIAM, 2)} mm")
print()

print(f"  {'PWM':>5}  {'Motor RPM':>9}  {'Wheel RPM':>9}  {'m/s':>6}  {'km/h':>6}  {'Torque mNm':>10}  {'Margin':>8}  {'Slip?':>22}  {'Rating'}")
print(f"  {SEP2}")

for p in PWM_POINTS:
    r = evaluate(MY_PINION_TEETH, MY_WHEEL_TEETH, p)
    print(f"  {p:>4}%  {r['motor_rpm']:>9,}  {r['wheel_rpm']:>9,}  {r['speed_ms']:>6.2f}  {r['speed_kmh']:>6.2f}  {r['wheel_torque']:>10.3f}  {r['t_margin']:>7.1f}%  {slip_label(r['slip']):>22}  {rating_label(r['t_margin'])}")

# Verdict on current ratio
r50      = evaluate(MY_PINION_TEETH, MY_WHEEL_TEETH, 50)
my_score = score_ratio(MY_PINION_TEETH, MY_WHEEL_TEETH)
print()
print(f"  SCORE FOR {MY_PINION_TEETH}:{MY_WHEEL_TEETH} = {my_score}/100")
print()

if r50["t_margin"] >= 40:
    verdict = "Your ratio is IDEAL for this robot. Good torque margin and useful speed range."
elif r50["t_margin"] >= 25:
    verdict = "Your ratio is ACCEPTABLE. Decent torque margin but little room for extra weight or fast cornering."
elif r50["t_margin"] >= 10:
    verdict = "Your ratio is MARGINAL. The robot moves but struggles under load. Consider a higher reduction."
elif r50["t_margin"] >= 0:
    verdict = "Your ratio is WEAK. Very little torque headroom — unreliable at low PWM. Increase reduction."
else:
    verdict = "Your ratio FAILS for this robot weight. Motor cannot overcome friction at 50% PWM. Increase reduction immediately."

print(f"  VERDICT: {verdict}")

# ---------- Ideal ratio finder ----------
print()
print(SEP)
print("  IDEAL RATIO FINDER — top candidates for your robot")
print(SEP)
print()

candidates = []
for p in PINION_RANGE:
    for w in WHEEL_RANGE:
        if math.gcd(p, w) != 1:
            continue   # hunting tooth principle: gcd must be 1 so every tooth pair eventually meshes
        s = score_ratio(p, w)
        if s < 0:
            continue
        r = evaluate(p, w, 50)
        candidates.append((s, p, w, r))

candidates.sort(key=lambda x: -x[0])
top = candidates[:8]

print(f"  {'Rank':<5}  {'Teeth':^9}  {'Ratio':>7}  {'Score':>7}  {'m/s @50%':>9}  {'Torque mNm':>10}  {'Margin':>8}  {'Center Dist':>12}  {'Wheel Diam':>10}")
print(f"  {SEP2}")

my_rank = None
for i, (s, p, w, r) in enumerate(top):
    marker = " <-- YOU" if (p == MY_PINION_TEETH and w == MY_WHEEL_TEETH) else ""
    if p == MY_PINION_TEETH and w == MY_WHEEL_TEETH:
        my_rank = i + 1
    print(f"  {i+1:<5}  {p}:{w:<6}  {r['ratio']:>7.2f}  {s:>7}  {r['speed_ms']:>9.2f}  {r['wheel_torque']:>10.3f}  {r['t_margin']:>7.1f}%  {r['c_dist']:>10.2f} mm  {r['w_diam']:>8.2f} mm{marker}")

if my_rank is None:
    full_rank = next((i+1 for i, (s,p,w,r) in enumerate(candidates) if p==MY_PINION_TEETH and w==MY_WHEEL_TEETH), None)
    if full_rank:
        print(f"\n  Your ratio {MY_PINION_TEETH}:{MY_WHEEL_TEETH} ranks #{full_rank} overall (outside top 8).")
    else:
        print(f"\n  Your ratio {MY_PINION_TEETH}:{MY_WHEEL_TEETH} was filtered out (check center distance or wheel size limits).")

# ---------- Best recommendation ----------
if top:
    best_s, best_p, best_w, best_r = top[0]
    print()
    print(SEP)
    print("  RECOMMENDATION")
    print(SEP)
    if my_rank == 1:
        print(f"\n  Your ratio {MY_PINION_TEETH}:{MY_WHEEL_TEETH} is already the top-ranked option for this robot.")
        print(f"  No changes needed.")
    else:
        print(f"\n  Best ratio for your robot: {best_p}:{best_w}  ({best_r['ratio']:.2f}:1)  Score: {best_s}/100")
        print(f"\n  {'Center distance':<28}: {best_r['c_dist']} mm")
        print(f"  {'Wheel diameter':<28}: {best_r['w_diam']} mm")
        print(f"  {'Speed at 50% PWM':<28}: {best_r['speed_ms']} m/s  ({best_r['speed_kmh']} km/h)")
        print(f"  {'Wheel torque at 50%':<28}: {best_r['wheel_torque']} mNm")
        print(f"  {'Torque margin at 50%':<28}: {best_r['t_margin']}%")
        if my_score >= 0:
            improvement = best_r['t_margin'] - r50['t_margin']
            print(f"\n  Your current ratio scores {my_score}/100 vs {best_s}/100 for the best option.")
            print(f"  Switching would give you {abs(round(improvement,1))}% {'more' if improvement > 0 else 'less'} torque margin.")

print()

# ================================================================
#  XLSX EXPORT (optional)
# ================================================================

XLSX_FILENAME = "coisa.xlsx"

if EXPORT_XLSX:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import ColorScaleRule

    wb = Workbook()
    ws = wb.active
    ws.title = "Ratio Results"

    # --- Header ---
    headers = ["rank", "pinion", "wheel", "score", "ratio",
               "motor_rpm", "wheel_rpm", "speed_ms", "speed_kmh",
               "wheel_torque", "req_torque", "t_margin", "current",
               "slip", "w_diam", "c_dist"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for i, (s, p, w, r) in enumerate(candidates, start=2):
        row = {"rank": i-1, "pinion": p, "wheel": w, "score": s, **r}
        for col, h in enumerate(headers, start=1):
            ws.cell(row=i, column=col, value=row[h])

    # --- Column widths ---
    col_widths = [6, 7, 7, 7, 7, 11, 11, 9, 9, 12, 11, 10, 9, 6, 9, 10]
    for col, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # --- Color scale on score column (D) ---
    ws.conditional_formatting.add(f"D2:D{len(candidates)+1}", ColorScaleRule(
        start_type="min", start_color="FF0000",
        end_type="max",   end_color="00FF00"
    ))

    wb.save(XLSX_FILENAME)
    print(f"  Results exported to {XLSX_FILENAME}  ({len(candidates)} combinations)")
    print()

if PLOT:
    fig, ax = plt.subplots(figsize=(10, 6))
    colors = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728"]  # one per PWM point

    for i, pct in enumerate(PWM_POINTS):
        r = evaluate(MY_PINION_TEETH, MY_WHEEL_TEETH, pct)

        stall_torque_wheel = MOTOR_STALL_TORQUE * (MY_WHEEL_TEETH/MY_PINION_TEETH) * GEAR_EFFICIENCY * (pct / 100)
        no_load_rpm_wheel  = (MOTOR_KV * MOTOR_VOLTAGE * (pct / 100)) / MY_RATIO
        x = [0, no_load_rpm_wheel]
        y = [stall_torque_wheel, 0]

        # T×n line
        ax.plot(x, y, color=colors[i], linewidth=2, label=f"{pct}% PWM")

        # Operating point — where load intersects the T×n line
        n_op = no_load_rpm_wheel * (1 - r["req_torque"] / stall_torque_wheel) if stall_torque_wheel > 0 else 0
        ax.scatter(n_op, r["req_torque"], color=colors[i], zorder=5, s=60)

    # Friction limit line — same for all PWM, grab from any evaluate() call
    r_ref = evaluate(MY_PINION_TEETH, MY_WHEEL_TEETH, 50)
    ax.axhline(y=r_ref["req_torque"], color="red", linestyle="--",
               linewidth=1.5, label=f"Min torque to move ({r_ref['req_torque']} mNm)")

    # Labels and formatting
    ax.set_title(f"Torque × Speed — {MY_PINION_TEETH}:{MY_WHEEL_TEETH}  ({MY_RATIO:.2f}:1)", fontsize=14)
    ax.set_xlabel("Wheel RPM")
    ax.set_ylabel("Wheel Torque (mNm)")
    ax.legend()
    ax.grid(True, linestyle="--", alpha=0.4)
    ax.set_xlim(left=0)
    ax.set_ylim(bottom=0)

    plt.tight_layout()
    plt.savefig("coisa_curve.png", dpi=150, bbox_inches="tight")
    plt.show()